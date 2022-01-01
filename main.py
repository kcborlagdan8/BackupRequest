from dns.rdatatype import NULL
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sftp import SFTP
import lookup
import sys

#Input main.py Cases.xlsx 1
file = sys.argv[1] #Cases.xlsx
type = sys.argv[2] #1

######################
#   Types of Cases   #
# 1 - Backup Request #
# 2 - Data Consent   #
# 3 - Support Pod    #
######################
def caseType(x):
    val = {
        "1": "database backup request",
        "2": "support server (not support pod)",
        "3": "support pod"
    }
    return val.get(x, "XXXXXX")

search = caseType(type) if type in ["1", "2", "3"] else exit("Choose from 1-3 only.")

wb = load_workbook(file)
ws = wb.active
count = 0 #number of tickets with type
dbtype = 'P'
dbserver = ''

def dbtypealias(dbtype):
    val = {
        "Preview": "D",
        "Production": "P",
        "Sandbox": "S"
    }
    return val.get(dbtype, "XXXXXX")

#loop through rows for backup requests
for row in range(1, 300): #row
    for col in range(1, 7): #column
        char = get_column_letter(col)
        val = ws[char + str(row)].value
        if(col == 6 and val is not None and search in str(val).lower()):
            count += 1
            
            #get values from excel
            clientID = ws["D" + str(row)].value
            client = ws["E" + str(row)].value
            case = ws["B" + str(row)].value
            product = ws["C" + str(row)].value

            #Generate sftp credentials
            try:
                session = SFTP().credGen(case)
                credentials = SFTP().getCred(session)
            except Exception:
                credentials = "SFTP already created"

            #Get instance name 
            domain = ws["G" + str(row)].value #get email
            instance = lookup.getInstanceName(domain)
            #end get instance name

            #Get database type
            subject = ws["F" + str(row)].value #check if preview, sandbox, production
            if(search == "database backup request"):
                dbtype = subject.split("-")
                dbtype = dbtype[1].strip()
                dbtype = dbtypealias(dbtype)
            #end database type           

            
            #get DB Server if NOT PREVIEW
            if(dbtype != "D"): 
                try:   
                    fqdn = lookup.dnsresolver(f"{instance}.deltekfirst.com") 
                except:
                    print(f"\nFQDN not found for {instance}, {clientID}. Check if the client has different instance name.")
                    fqdn = ""

                if(product == "Vision"):
                    dbserver = lookup.getDBfromCNAMEDFVE(fqdn)
                else:
                    dbserver = lookup.getVTPod(fqdn)
            else:
                dbserver = "USEAPVVT0DB1"


            #get database name for VT
            databaseName = instance #default
            
            if(product == "Vantagepoint" or dbtype == "D"):
                databaseName = "C00000" \
                    + clientID \
                    + dbtype \
                    + "_1_" \
                    + ((instance.upper() + "00000000000")[:11]) #e.g.C000073333P_1_ontoit00000

            if(dbtype == "S" and product == "Vision"):
                databaseName += "_Sandbox"
    
            #print(databaseName)
            cred = ",".join(credentials)
            print(f"{dbserver},{databaseName},{clientID},N,Y,\"{client}\",{case},{cred}")

            
print(f"\nNumber of Backup Requests: {str(count)}")