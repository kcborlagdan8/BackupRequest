from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
from Case import *
from sftp import SFTP
import lookup
import sys
import secrets


#Input main.py Cases.xlsx 1
file = sys.argv[1] #Cases.xlsx
type = sys.argv[2] #1

######################
#   Types of Cases   #
# 1 - Backup Request #
# 2 - Data Consent   #
# 3 - Support Pod
# 4 - Database Refresh (Preview NOT included)
# 9 - All SRs        #
# 0 - All Incidents  #
######################

#Types of Case
dbr = "database backup request"
def caseType(x):
    cType = {
        "1": ["database backup request","database backup download request","[ajera cloud support] requesting database:"],
        "2": ["support server (not support pod)", "local restore", "local server", "backup copy", "local ftp"],
        "3": ["support pod"],
        "4": ["database refresh - refresh sandbox from a copy of production", "database refresh - refresh production from a copy of sandbox"],
        "5": ["database refresh - refresh preview from a copy of productio"]
    }
    #All SRs "9"
    cType["9"] = [cType["1"][0], cType["4"][0], cType["5"][0]]

    #All incidents "0"
    cType["0"] = [cType["2"][0], cType["2"][1], cType["3"][0]]

    return cType.get(x, "XXXXXX")
 
search = caseType(type) if type in ["1", "2", "3", "4", "5", "9", "0"] else exit("Choose from 0-9 only.")

wb = load_workbook(file)
ws = wb.active
count = 0 #number of tickets with case type
brcount = 0
dbrcount = 0
#dccount = 0
spcount = 0

brline = ""
dbrline = ""
#dcline = ""
spline = ""
ajeraline = ""

notASCline = ""
s = ", ".join(search).capitalize()
print(f"Search for: {s}")

#LOOP THROUGH EXCEL FILE
#print(type, search)
for row in range(1, 300): #row
    for col in range(1, 7): #column
        char = get_column_letter(col)
        val = ws[char + str(row)].value
        if(col == 6 and val is not None and any(y in str(val).lower() for y in search)):
            count += 1
            #create dict from excel
            request = {
                "clientID": ws["D" + str(row)].value,
                "client": ws["E" + str(row)].value,
                "rnt": ws["B" + str(row)].value,
                "email": ws["G" + str(row)].value,
                "dlz": ws["H" + str(row)].value,
                "product": ws["C" + str(row)].value,
                "subject": ws["F" + str(row)].value,
                "dataconsentdate": ws["I" + str(row)].value,
                "asc": ws["J" + str(row)].value,
            }
            
            if(type in ["1", "2", "9"]# and val.lower().find(search[0]) != -1
            ): #Backup Request
                if(request['asc'] != 'Yes'):
                    notASCline += request['rnt']     
                    continue
                br = BackupRequest(request)                    
                br.generateSFTP()
                cred = ",".join(br.credentials) if (br.credentials != "SFTP credentials already created or SFTP site is inaccessible") else br.credentials
                br.rnt = br.rnt.replace("-","")
                #Ajera database backup download request
                if(request["product"] == "Ajera" and type in ["1","2"]):
                    ajeradbs = br.ajera()
                    for a in ajeradbs:
                        
                        ajeraline += f"""{a[0]},{a[1]},{a[2]},N,Y,"{a[3]}",RNT{br.rnt},{br.dataconsentdate},{cred}
"""#,{cred},{secrets.token_urlsafe(10)},{secrets.token_urlsafe(10)}                     
                    brline += ajeraline
                    continue

                elif(val.lower().find(search[0]) != -1 and type == "1"):
                    dbtype = br.getDBType()
                    br.getDB(dbtype)
                
                else: #data consent 
                    br.getDB("P")
                    

                if(request["product"] != "Ajera"): brline +=f"""{br.dbserver},{br.db},{br.clientID},N,Y,\"{br.client}\",RNT{br.rnt},{br.dataconsentdate},{cred}
""" #,{secrets.token_urlsafe(10)},{secrets.token_urlsafe(10)}
                brcount += 1
            
            if(request["product"] != "Ajera" and type == "4" or type == "5" or type == "9" and (val.lower().find(search[1]) != -1 or val.lower().find(search[2]) != -1)): #DB Refresh
                dbr = DatabaseRefresh(request)
                dbr.getDBs()
                
                if(dbr.dbserver == "XXXXXX"): #if fqdn was not found
                    dbr.sourceDB = dbr.dbserver 
                    dbr.destDB = dbr.dbserver 

                if(type == "5"):
                    dbrline +=f"""{dbr.dbserver},{dbr.sourceDB},{dbr.clientID},<get Preview AD account from VT PII>
"""
                    dbrline +=f"""{dbr.dbserver},{dbr.sourceDB}Files,{dbr.clientID},<get Preview AD account from VT PII>
"""                   
                else:
                    dbrline +=f"""{dbr.dbserver},{dbr.clientID},{dbr.instance},{dbr.sourceDB},{dbr.destDB},{dbr.rnt}
"""
                dbrcount += 1
            # if(cType == "9"):
            #     line = """
            #     """
print(f"\nBackup and Data Consent: {brcount}")               
print(f"\033[32m{brline}\033[0m")
print(f"Database Refresh: {dbrcount}")               
print(dbrline)
#print("Data Consent:")               
#print(dcline)
print("Support Pod:")               
print(spline)

print("Cases with Not Authorized Contact(did not process):")
print(f"\033[31m{notASCline}\033[0m")

print(f"\nNumber of cases to be processed: {str(count)}")